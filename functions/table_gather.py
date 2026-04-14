import os
import openpyxl


tgt_langs = ["zh", "lo", "ja", "vi", "id", "fr", "es", "si", "de"]
METRICS = ["spbleu", "chrf", "bleurt",
           "comet", "xcomet", "reg",
           "kiwi", "kiwi23", "qe"]
numbers = [1, 2, 3, 4, 5]

output_file = "excel/all.xlsx"

output_wb = None
output_ws = None

if os.path.exists(output_file):
    os.remove(output_file)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "sheet1"
for i, metric in enumerate(METRICS):
    metric_col = 2 + i * (len(numbers) + 1)
    sheet.cell(row=1, column=metric_col).value = metric
    for count, number in enumerate(numbers, start=0):
        sheet.cell(row=2, column=metric_col + count).value = number

for tgt_lang in tgt_langs: 
    tmp_path = f"excel/{tgt_lang}.xlsx"
    if not os.path.exists(tmp_path):
        sheet.append([f"en-{tgt_lang}"])
    else:
        wb_tmp = openpyxl.load_workbook(tmp_path)
        ws = wb_tmp.active
        third_row = [cell.value for cell in ws[3]]
        sheet.append(third_row)

wb.save(output_file)
wb.close()
