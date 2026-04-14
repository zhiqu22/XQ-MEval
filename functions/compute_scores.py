import score_tools as tool
import pandas as pd
from pathlib import Path
import sys
import openpyxl
import os
from statistics import mean, variance

METRICS = ["spbleu", "chrf", "bleurt",
           "comet", "xcomet", "reg",
           "kiwi", "kiwi23", "qe"]

tgt_lang = sys.argv[1]
metric = sys.argv[2]
gpu_nums = int(sys.argv[3])
batch_size = int(sys.argv[4])
value_dict = {}
results_dict = {}
numbers = [1, 2, 3, 4, 5]
iterations = range(1, 11)

def make_and_add_table(tgt_lang, metric, value_dict):
    file_path = os.path.join("./excel", f"{tgt_lang}.xlsx")
    
    if not os.path.exists("./excel"):
        os.makedirs("./excel")

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "sheet1"
        sheet.cell(row=3, column=1).value = f"en-{tgt_lang}"

    metric_col = None
    for i, tmp in enumerate(METRICS):
        if tmp == metric:
            metric_col = 2 + i * (len(numbers) + 1)
            sheet.cell(row=1, column=metric_col).value = metric
            break

    if metric_col:
        for count, number in enumerate(numbers, start=0):
            sheet.cell(row=2, column=metric_col + count).value = number
            sheet.cell(row=3, column=metric_col + count).value = value_dict.get(f"{metric}-{number}", "N/A")

    wb.save(file_path)
    wb.close()

def read_data(tgt_lang, number,iteration):
    current_dir = Path.cwd()
    file_path = os.path.join(current_dir, "merged_result", f"en-{tgt_lang}-merge-{number}_{iteration}.tsv")
    raw_data = pd.read_csv(file_path, sep="\t")
    srcs, hyps, refs = [], [], []
    for _, row in raw_data.iterrows():
        srcs.append(row["src"])
        hyps.append(row["merged_mt"])
        refs.append(row["ref"])
    assert srcs is not None
    assert hyps is not None
    assert refs is not None
    return srcs, hyps, refs

if metric == "chrf" or metric == "spbleu":
    for number in numbers:
        scores = []
        for iteration in iterations:
          srcs, hyps, refs = read_data(tgt_lang, number, iteration)
          input_data = tool.prepare_input_data(None, hyps, refs, metric_type="seq", metric_name=metric)
          results = tool.seq_score(input_data, metric)
          results_dict[f"{metric}-{number}_{iteration}"] = results.system_score
          score = results.system_score
          scores.append(score)
        mean_score = round(mean(scores), 6)
        value_dict[f"{metric}-{number}"]= mean_score

    
elif metric == "bleurt":

    for number in numbers:
        multiple_hyps, multiple_refs = [], []
        scores = []
        for iteration in iterations:
            _, hyps, refs = read_data(tgt_lang, number,iteration)
            multiple_hyps.append(hyps)
            multiple_refs.append(refs)
        input_data, segments = tool.prepare_batch_bleurt_input_data(multiple_hyps,multiple_refs)
        multiple_results = tool.bleurt_score(input_data, segments)
        assert len(multiple_results)  == len(iterations)
        for i in range(len(multiple_results)):
            iteration = iterations[i]
            results_dict[f"{metric}-{number}_{iteration}"] = multiple_results[i].system_score*100 
            score = multiple_results[i].system_score*100
            scores.append(score)
        mean_score = round(mean(scores), 6)
        value_dict[f"{metric}-{number}"]= mean_score


elif metric in ["kiwi23", "kiwi", "comet", "xcomet"]:
    for number in numbers:
        multiple_pairs =[]
        scores = []
        for iteration in iterations:
            
            srcs, hyps, refs = read_data(tgt_lang, number, iteration)

            if metric == "kiwi23" or metric == "kiwi":
                input_data = tool.prepare_input_data(src=srcs, hyp=hyps, ref=None, metric_type="qe", metric_name=metric)
            else:
                input_data = tool.prepare_input_data(src=srcs, hyp=hyps, ref=refs, metric_type="reg", metric_name=metric)
            multiple_pairs.append(input_data)
        multiple_results = tool.comet_score(multiple_pairs, metric, batch_size=batch_size, gpus=gpu_nums)
        assert len(multiple_results) == len(iterations)   
        for i in range(len(multiple_results)):
            iteration = iterations[i]
            results_dict[f"{metric}-{number}_{iteration}"] = multiple_results[i].system_score*100
            score = multiple_results[i].system_score*100
            scores.append(score)
        mean_score = round(mean(scores), 6)
        value_dict[f"{metric}-{number}"]= mean_score



elif metric == "reg":
    for number in numbers:
        multiple_hyps, multiple_refs = [], []
        scores = []
        for iteration in iterations:
            _, hyps, refs = read_data(tgt_lang, number, iteration)
            multiple_hyps.append(hyps)
            multiple_refs.append(refs)
        dir_path, segments = tool.prepare_batch_metricx_input_data(None, multiple_hyps, multiple_refs, metric_type="reg")
        multiple_results = tool.metricx_score(dir_path, metric_type="reg", segments=segments, batch_size=1)
        assert len(multiple_results) == len(iterations)
        for i in range(len(multiple_results)):
            iteration = iterations[i]
            results_dict[f"{metric}-{number}_{iteration}"] = multiple_results[i].system_score
            score = multiple_results[i].system_score
            scores.append(score)
        mean_score = round(mean(scores), 6)
        value_dict[f"{metric}-{number}"]= mean_score
        


elif metric == "qe":
    for number in numbers:
        multiple_hyps, multiple_srcs = [], []
        scores = []
        for iteration in iterations:
            srcs, hyps, refs = read_data(tgt_lang, number, iteration)
            multiple_hyps.append(hyps)
            multiple_srcs.append(srcs)
        dir_path, segments = tool.prepare_batch_metricx_input_data(multiple_srcs, multiple_hyps, None, metric_type="qe")
        multiple_results = tool.metricx_score(dir_path, metric_type="qe", segments=segments, batch_size=1)
        assert len(multiple_results) == len(iterations)
        for i in range(len(multiple_results)):
            iteration = iterations[i]
            results_dict[f"{metric}-{number}_{iteration}"] = multiple_results[i].system_score
            score = multiple_results[i].system_score
            scores.append(score)
        mean_score = round(mean(scores), 6)
        value_dict[f"{metric}-{number}"] = mean_score

print(results_dict)
print(value_dict)
make_and_add_table(tgt_lang, metric, value_dict)
